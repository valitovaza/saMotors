from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import UTC, datetime
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config.workbook_schema import EXPORT_ROOT, LOCAL_DB, LOCAL_WORKBOOK, SUPPORTED_SHEETS
from app.db import connect_sqlite
from app.services.state_service import build_app_state, load_vehicle_rows


MANAGED_EXPORT_SHEETS = (
    "Front Sheet",
    "Stock Data",
    "Sold Stock",
    "Collection",
    "Investor Budget",
    "Expense",
    "Fuel Expense",
    "Investor Car Expense",
    "Money in",
    "Cash Spending",
    "Money Out",
)


@dataclass
class ExportSummary:
    workbook_path: str
    filename: str
    sync_run_id: int | None
    exported_sheets: list[str]
    untouched_sheets: list[str]
    rows_written: dict[str, int]


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _parse_iso_date(value: str | None) -> datetime | None:
    text = _clean(value)
    if not text:
        return None
    for candidate in (text, f"{text}T00:00:00"):
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00"))
        except ValueError:
            continue
    return None


def _month_bucket(*values: str | None) -> str:
    for value in values:
        parsed = _parse_iso_date(value)
        if parsed is not None:
            return parsed.strftime("%Y-%m-01")
    return ""


def _detect_header_row(worksheet: Worksheet, expected_headers: list[str]) -> int | None:
    expected = {_clean(header) for header in expected_headers}
    best_row = None
    best_matches = 0
    for row_index in range(1, min(worksheet.max_row, 10) + 1):
        row_values = {_clean(worksheet.cell(row_index, column).value) for column in range(1, worksheet.max_column + 1)}
        row_values.discard("")
        matches = len(expected & row_values)
        if matches > best_matches:
            best_matches = matches
            best_row = row_index
    return best_row if best_matches >= 2 else None


def _ensure_headers(workbook: Workbook, sheet_name: str, expected_headers: list[str]) -> tuple[Worksheet, int, dict[str, int]]:
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    header_row = _detect_header_row(worksheet, expected_headers)
    if header_row is None:
        header_row = 1
        for column_index, header in enumerate(expected_headers, start=1):
            worksheet.cell(header_row, column_index, header)

    header_map: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header = _clean(worksheet.cell(header_row, column_index).value)
        if header:
            header_map[header] = column_index

    next_column = max(header_map.values(), default=0) + 1
    for header in expected_headers:
        if header not in header_map:
            worksheet.cell(header_row, next_column, header)
            header_map[header] = next_column
            next_column += 1

    return worksheet, header_row, header_map


def _replace_sheet_rows(worksheet: Worksheet, header_row: int, header_map: dict[str, int], rows: list[dict[str, Any]]) -> None:
    if worksheet.max_row > header_row:
        worksheet.delete_rows(header_row + 1, worksheet.max_row - header_row)

    ordered_headers = sorted(header_map.items(), key=lambda item: item[1])
    for row_offset, payload in enumerate(rows, start=1):
        target_row = header_row + row_offset
        for header, column_index in ordered_headers:
            worksheet.cell(target_row, column_index, payload.get(header, ""))


def _vehicle_sheet_row(vehicle: dict[str, Any], sold: bool) -> dict[str, Any]:
    month_value = _month_bucket(vehicle.get("date_sold"), vehicle.get("date_acquired")) or vehicle.get("date_acquired") or ""
    base_row = {
        "Month": month_value,
        "Date Aquired": vehicle.get("date_acquired") or "",
        "Make & Model": vehicle.get("display_name") or vehicle.get("model") or "",
        "Total Cost": round(float(vehicle.get("total_cost_cached") or 0), 2),
        "Sold": round(float(vehicle.get("sold_price") or 0), 2),
        "Stock ID": vehicle.get("stock_id") or "",
    }
    if sold:
        base_row.update(
            {
                "Number Plate reference": vehicle.get("plate") or "",
                "SA/Investor Name": vehicle.get("investor_name") or "SA",
                "Part Ex": round(float(vehicle.get("px_value") or 0), 2),
                "SA/Investor Profit Share": round(float(vehicle.get("profit_share_percent") or 0), 2),
                "Total Profit": round(float(vehicle.get("profit_total_cached") or 0), 2),
                "Investor Profit": round(float(vehicle.get("investor_profit_total") or 0), 2),
                "SA Profit": round(float(vehicle.get("company_profit_total") or 0), 2),
                "Date Listed": vehicle.get("date_listed") or "",
                "Date Sold": vehicle.get("date_sold") or "",
                "Days to Sell": max(
                    0,
                    ((
                        _parse_iso_date(vehicle.get("date_sold"))
                        and _parse_iso_date(vehicle.get("date_acquired"))
                    ) and (
                        _parse_iso_date(vehicle.get("date_sold")).date() - _parse_iso_date(vehicle.get("date_acquired")).date()
                    ).days)
                    or 0,
                ),
                "Platfrom": vehicle.get("platform") or "",
                "Invoice Number": vehicle.get("invoice_number") or "",
                "Customer Name": vehicle.get("customer_name") or "",
                "Contact info": vehicle.get("contact_info") or "",
                "Warranty": vehicle.get("warranty") or "",
                "AutoGuard Number": vehicle.get("autoguard_number") or "",
            }
        )
    else:
        base_row.update(
            {
                "Plate Number": vehicle.get("plate") or "",
                "Investor/SA": vehicle.get("investor_name") or "SA",
                "Source": vehicle.get("source") or "",
                "PX Value": round(float(vehicle.get("px_value") or 0), 2),
                "Price": round(float(vehicle.get("purchase_price") or 0), 2),
                "Reconditioning costs": round(float(vehicle.get("reconditioning_costs") or 0), 2),
                "Profit": round(float(vehicle.get("profit_total_cached") or 0), 2),
                "Status": vehicle.get("status") or "In Stock",
            }
        )
    return base_row


def _load_investors(connection: Any) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT *
        FROM investors
        ORDER BY name COLLATE NOCASE ASC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def _load_collections(connection: Any) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            cd.*,
            v.stock_id,
            v.plate,
            v.display_name
        FROM collections_deliveries cd
        LEFT JOIN vehicles v ON v.id = cd.vehicle_id
        WHERE lower(cd.job_type) = 'collection'
        ORDER BY COALESCE(cd.scheduled_date, cd.date_won, cd.created_at) DESC, cd.id DESC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def _load_vehicle_expenses(connection: Any) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            ve.*,
            v.stock_id,
            v.plate,
            v.display_name
        FROM vehicle_expenses ve
        JOIN vehicles v ON v.id = ve.vehicle_id
        ORDER BY COALESCE(ve.expense_date, ve.created_at) DESC, ve.id DESC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def _load_money_movements(connection: Any) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            mm.*,
            v.stock_id,
            v.plate,
            v.display_name
        FROM money_movements mm
        LEFT JOIN vehicles v ON v.id = mm.vehicle_id
        ORDER BY COALESCE(mm.movement_date, mm.created_at) DESC, mm.id DESC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def _expense_sheet_rows(expenses: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    expense_rows: list[dict[str, Any]] = []
    fuel_rows: list[dict[str, Any]] = []
    investor_car_rows: list[dict[str, Any]] = []

    for row in expenses:
        month_value = _month_bucket(row.get("expense_date")) or row.get("expense_date") or ""
        source_sheet = _clean(row.get("source_sheet"))
        amount = round(float(row.get("amount") or 0), 2)

        if source_sheet == "Investor Car Expense" or row.get("expense_type") == "investor_car_expense":
            investor_car_rows.append(
                {
                    "Month": month_value,
                    "Date": row.get("expense_date") or "",
                    "Reason": row.get("description") or row.get("category") or "",
                    "Amount": amount,
                    "Reg": row.get("plate") or "",
                    "Stock ID": row.get("stock_id") or "",
                }
            )
            continue

        if source_sheet == "Fuel Expense" or _clean(row.get("category")).lower() == "fuel" or row.get("expense_type") == "fuel_expense":
            fuel_rows.append(
                {
                    "Month": month_value,
                    "Date": row.get("expense_date") or "",
                    "Car": row.get("plate") or row.get("display_name") or "",
                    "Amount ": amount,
                    "Stock ID": row.get("stock_id") or "",
                }
            )
            continue

        expense_rows.append(
            {
                "Month": month_value,
                "Date": row.get("expense_date") or "",
                "Category": row.get("category") or "Other",
                "From": row.get("vendor") or "",
                "Amount ": amount,
                "Payment Method": row.get("payment_method") or "",
                "Paid By": row.get("paid_by") or "",
                "Notes": row.get("notes") or row.get("description") or "",
                "Stock ID": row.get("stock_id") or "",
            }
        )

    return expense_rows, fuel_rows, investor_car_rows


def _movement_sheet_rows(movements: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    money_in_rows: list[dict[str, Any]] = []
    cash_spending_rows: list[dict[str, Any]] = []
    money_out_rows: list[dict[str, Any]] = []

    for row in movements:
        month_value = _month_bucket(row.get("movement_date")) or row.get("movement_date") or ""
        payload = {
            "Month": month_value,
            "Date": row.get("movement_date") or "",
            "Category": row.get("category") or "",
            "Amount ": round(float(row.get("amount") or 0), 2),
            "Notes": row.get("notes") or "",
            "Stock ID": row.get("stock_id") or "",
        }
        source_sheet = _clean(row.get("source_sheet"))
        direction = _clean(row.get("direction")).lower()

        if source_sheet == "Cash Spending":
            cash_spending_rows.append(
                {
                    "Month": month_value,
                    "Amount ": payload["Amount "],
                    "Cost Incurred on": row.get("plate") or row.get("stock_id") or row.get("counterparty") or "",
                    "Reason": row.get("notes") or row.get("category") or "",
                    "Stock ID": row.get("stock_id") or "",
                }
            )
            continue

        if source_sheet == "Money in" or direction == "in":
            money_in_rows.append(
                {
                    "Month": month_value,
                    "Date": payload["Date"],
                    "Category": payload["Category"],
                    "Amount ": payload["Amount "],
                    "Reg": row.get("plate") or "",
                    "Notes": payload["Notes"],
                    "Stock ID": row.get("stock_id") or "",
                }
            )
            continue

        money_out_rows.append(payload)

    return money_in_rows, cash_spending_rows, money_out_rows


def _build_export_rows(db_path: Path) -> dict[str, list[dict[str, Any]]]:
    vehicle_rows = load_vehicle_rows(db_path)
    sold_rows = [row for row in vehicle_rows if _clean(row.get("status")).lower() == "sold" or row.get("date_sold")]
    stock_rows = [row for row in vehicle_rows if row not in sold_rows]

    with connect_sqlite(db_path) as connection:
        investors = _load_investors(connection)
        collections = _load_collections(connection)
        vehicle_expenses = _load_vehicle_expenses(connection)
        money_movements = _load_money_movements(connection)

    expense_rows, fuel_rows, investor_car_rows = _expense_sheet_rows(vehicle_expenses)
    money_in_rows, cash_spending_rows, money_out_rows = _movement_sheet_rows(money_movements)
    app_state = build_app_state(db_path)

    sold_by_month: dict[str, dict[str, float]] = {}
    for sold in app_state.get("sold", []):
        month = sold.get("month") or _month_bucket(sold.get("date_sold"), sold.get("date_acquired"))
        if not month:
            continue
        bucket = sold_by_month.setdefault(
            month,
            {
                "investor_profit": 0.0,
                "company_profit": 0.0,
            },
        )
        bucket["investor_profit"] += round(float(sold.get("investor_profit") or 0), 2)
        bucket["company_profit"] += round(float(sold.get("mp_profit") or 0), 2)

    investor_expense_by_month: dict[str, float] = {}
    for row in investor_car_rows:
        month = row.get("Month") or ""
        if not month:
            continue
        investor_expense_by_month[month] = investor_expense_by_month.get(month, 0.0) + round(float(row.get("Amount") or 0), 2)

    other_money_in_by_month: dict[str, float] = {}
    for row in money_in_rows:
        month = row.get("Month") or ""
        if not month:
            continue
        other_money_in_by_month[month] = other_money_in_by_month.get(month, 0.0) + round(float(row.get("Amount ") or 0), 2)

    other_money_out_by_month: dict[str, float] = {}
    for row in money_out_rows + cash_spending_rows:
        month = row.get("Month") or ""
        if not month:
            continue
        amount_key = "Amount " if "Amount " in row else "Amount"
        other_money_out_by_month[month] = other_money_out_by_month.get(month, 0.0) + round(float(row.get(amount_key) or 0), 2)

    front_sheet_rows = []
    for month_row in app_state.get("monthly", []):
        month = month_row.get("month") or ""
        profits = sold_by_month.get(month, {})
        company_expenses = round(float(month_row.get("expenses") or 0), 2)
        company_fuel = round(float(month_row.get("fuel") or 0), 2)
        other_money_in = round(other_money_in_by_month.get(month, 0.0), 2)
        other_money_out = round(other_money_out_by_month.get(month, 0.0), 2)
        total_sa_gross_profit = round(profits.get("company_profit", 0.0), 2)
        investor_net_profit = round(profits.get("investor_profit", 0.0), 2)
        investor_expense = round(investor_expense_by_month.get(month, 0.0), 2)
        net_profit_exc_investor = round(total_sa_gross_profit - company_expenses - company_fuel + other_money_in - other_money_out, 2)
        front_sheet_rows.append(
            {
                "Month": month,
                "Cars Sold": int(month_row.get("cars_sold") or 0),
                "Total Revenue": round(float(month_row.get("revenue") or 0), 2),
                "Total Gross Profit": round(float(month_row.get("gross_profit") or 0), 2),
                "Company Expenses": company_expenses,
                "Total SA Gross Profit": total_sa_gross_profit,
                "Investor Net Profit": investor_net_profit,
                "Investor Expense": investor_expense,
                "Company Fuel Costs": company_fuel,
                "Other Money In": other_money_in,
                "Other Money Out": other_money_out,
                "Net Profit Exc Investor": net_profit_exc_investor,
                "Net Exc Investor": net_profit_exc_investor,
                "Notes": "",
            }
        )

    return {
        "Front Sheet": front_sheet_rows,
        "Stock Data": [_vehicle_sheet_row(row, sold=False) for row in stock_rows],
        "Sold Stock": [_vehicle_sheet_row(row, sold=True) for row in sold_rows],
        "Collection": [
            {
                "Source": row.get("source") or "",
                "Date Won": row.get("date_won") or "",
                "Plate Number": row.get("plate") or "",
                "Make & Model": row.get("display_name") or "",
                "Location": row.get("address") or "",
                "Post Code": row.get("postcode") or "",
                "How Far?": row.get("distance_note") or "",
                "Collection Date": row.get("scheduled_date") or "",
                "Number": row.get("contact_number") or "",
                "Additional notes": row.get("notes") or "",
                "Stock ID": row.get("stock_id") or "",
            }
            for row in collections
        ],
        "Investor Budget": [
            {
                "Investors": row.get("name") or "",
                "Initial Balance": round(float(row.get("initial_balance") or 0), 2),
                "Capital Returned": round(float(row.get("capital_returned") or 0), 2),
                "Total Balance": round(float(row.get("total_balance_cached") or 0), 2),
                "Purchased": round(float(row.get("purchased_total_cached") or 0), 2),
                "Total Profit (since Nov-25)": round(float(row.get("profit_total_cached") or 0), 2),
                "Available": round(float(row.get("available_balance_cached") or 0), 2),
            }
            for row in investors
        ],
        "Expense": expense_rows,
        "Fuel Expense": fuel_rows,
        "Investor Car Expense": investor_car_rows,
        "Money in": money_in_rows,
        "Cash Spending": cash_spending_rows,
        "Money Out": money_out_rows,
    }


def export_database_to_workbook(db_path: Path = LOCAL_DB, template_workbook_path: Path = LOCAL_WORKBOOK) -> ExportSummary:
    if not template_workbook_path.exists():
        raise ValueError("Local workbook template does not exist.")

    EXPORT_ROOT.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    export_filename = f"DealerOS_export_{timestamp}.xlsx"
    export_path = EXPORT_ROOT / export_filename

    with connect_sqlite(db_path) as connection:
        sync_run_id = connection.execute(
            """
            INSERT INTO workbook_sync_runs (direction, status, source_path)
            VALUES ('export', 'running', ?)
            """,
            (str(template_workbook_path),),
        ).lastrowid
        connection.commit()

    workbook = load_workbook(template_workbook_path)
    export_rows = _build_export_rows(db_path)
    rows_written: dict[str, int] = {}
    exported_sheets: list[str] = []

    for sheet_name in MANAGED_EXPORT_SHEETS:
        headers = SUPPORTED_SHEETS[sheet_name]
        worksheet, header_row, header_map = _ensure_headers(workbook, sheet_name, headers)
        rows = export_rows.get(sheet_name, [])
        _replace_sheet_rows(worksheet, header_row, header_map, rows)
        rows_written[sheet_name] = len(rows)
        exported_sheets.append(sheet_name)

    workbook.save(export_path)

    untouched_sheets = [sheet for sheet in workbook.sheetnames if sheet not in exported_sheets]
    summary = ExportSummary(
        workbook_path=str(export_path),
        filename=export_filename,
        sync_run_id=int(sync_run_id),
        exported_sheets=exported_sheets,
        untouched_sheets=untouched_sheets,
        rows_written=rows_written,
    )

    with connect_sqlite(db_path) as connection:
        connection.execute(
            """
            UPDATE workbook_sync_runs
            SET status = 'completed',
                finished_at = CURRENT_TIMESTAMP,
                summary_json = ?,
                conflicts_json = ?
            WHERE id = ?
            """,
            (
                json.dumps(asdict(summary)),
                json.dumps({"untouched_sheets": untouched_sheets}),
                sync_run_id,
            ),
        )
        connection.commit()

    return summary
